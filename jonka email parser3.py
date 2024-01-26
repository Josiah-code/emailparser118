import os
import imaplib
import re
from email import message_from_bytes
from openpyxl import Workbook
from bs4 import BeautifulSoup
from datetime import datetime

# Function to extract location from the email body between specified phrases
def extract_location(email_body):
    location_start = email_body.find("Role request Review the request and respond.")
    location_end = email_body.find("is requesting to be listed")
    
    if location_start != -1 and location_end != -1:
        location = email_body[location_start + len("Role request Review the request and respond."):location_end].strip()

        # Replace '\u200d' with blank spaces and remove extra spaces
        location = location.replace('\u200d', ' ').strip()
        location = " ".join(location.split())  # Clean up extra spaces
        
        return location
    else:
        return "Location Not Found"


# Function to extract the line that starts with '<https://business.google.com/n/' and ends with '>'
def extract_business_link_line(email_body):
    link_line = re.search(r'<https://business.google.com/n/[^>]+>', email_body)
    if link_line:
        return link_line.group()
    else:
        return "Business Link Line Not Found"

# Function to extract business name from emails with a modified subject
def extract_business_name(email_body):
    business_name_match = re.search(r'as (?:an owner|a manager) of (.*?) Business Profile on Google', email_body)
    if business_name_match:
        return business_name_match.group(1).strip()
    else:
        return "Business Name Not Found"

# Function to extract business name and business link from emails with a modified subject
def extract_suspended_business_info(email_body):
    business_info_start = email_body.find("Your Business Profile has been suspended")
    business_info_end = email_body.find("your Business Profile on Google has been suspended because it was flagged for suspicious activity")

    if business_info_start != -1 and business_info_end != -1:
        business_info = email_body[business_info_start + len("Your Business Profile has been suspended"):business_info_end].strip()
        business_name = re.search(r'^(.*?)<', business_info).group(1).strip() if re.search(r'^(.*?)<', business_info) else "Business Name Not Found"
        business_link = re.search(r'<(https://business.google.com/n/[^>]+) > at', business_info).group(1).strip() if re.search(r'<(https://business.google.com/n/[^>]+) > at', business_info) else "Business Link Not Found"
        return business_name, business_link

    return "Business Name Not Found", "Business Link Not Found"

# Function to extract email address from email body
def extract_email_address(email_body):
    email_match = re.search(r'Email\s*address\s*:\s*([\w\.-]+@[\w\.-]+)', email_body, re.IGNORECASE)
    if email_match:
        return email_match.group(1).strip()
    else:
        return "Email Address Not Found"

# Function to extract business name from emails with the subject "Your post has been removed from Google"
def extract_removed_post_business_name(email_body):
    business_name_match = re.search(r'(.*?) your post has been removed from your Business Profile on Google because it contains content that is considered spam.', email_body, re.DOTALL)
    if business_name_match:
        return business_name_match.group(1).strip()
    else:
        return "Business Name Not Found"

# Prompt the user to choose between parsing all emails or only unread ones
parse_all = input("Parse all emails (y/n)? ").strip().lower() == "y"

if parse_all:
    # Define the date range for email filtering
    start_date = datetime(2023, 11, 8)  # Replace with your desired start date
    end_date = datetime(2023, 11, 14)  # Replace with your desired end date
else:
    # If not parsing all emails, prompt the user for the start and end dates
    start_date_str = input("Enter the start date (YYYY-MM-DD): ")
    end_date_str = input("Enter the end date (YYYY-MM-DD): ")
    start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
    end_date = datetime.strptime(end_date_str, "%Y-%m-%d")

# Convert the dates to the required format (DD-Mon-YYYY)
start_date_str = start_date.strftime("%d-%b-%Y")
end_date_str = end_date.strftime("%d-%b-%Y")

# Connect to your email server
mail = imaplib.IMAP4_SSL("md-uk-1.webhostbox.net")
mail.login("parser@taskmoby.com", "Welcome@118")
mail.select("inbox")

# Modify the search criteria to filter emails within the date range
search_criteria = f'(SENTSINCE {start_date_str} SENTBEFORE {end_date_str})'

# If parsing only unread emails, add the UNSEEN criterion
if not parse_all:
    search_criteria += 'SEEN'

result, data = mail.search(None, search_criteria)
email_ids = data[0].split()

# Create a new workbook for suspended emails
suspended_workbook = Workbook()
suspended_sheet = suspended_workbook.active
suspended_sheet.title = "Suspended Emails"
suspended_sheet.append(['Business Name', 'Business Link Line', 'Email Content'])

# Create the main workbook
workbook = Workbook()
sheet = workbook.active
sheet.append(['Business Name', 'Location', 'Business Link Line', 'Name', 'Phone Number', 'Email', 'Date and Time Sent'])

# Create the "suspended_posts" workbook
suspended_posts_workbook = Workbook()
suspended_posts_sheet = suspended_posts_workbook.active
suspended_posts_sheet.title = "Suspended Posts"
suspended_posts_sheet.append(['Business Name', 'Email Content'])

for email_id in email_ids:
    try:
        result, data = mail.fetch(email_id, '(RFC822)')
        msg = message_from_bytes(data[0][1])

        subject = msg.get("subject", "")
        print(f"Processing email with subject: {subject}")

        if "You=E2=80=99ve_received_a_management_request" in subject or "You=E2=80=99ve_received_an_ownership_request" in subject:
            # Process Management or Ownership Request emails
            content_type = msg.get_content_type()
            if content_type == "multipart/alternative":
                email_body = None
                for part in msg.walk():
                    part_content_type = part.get_content_type()
                    if part_content_type == "text/plain" and email_body is None:
                        email_body = part.get_payload(decode=True).decode()
                    elif part_content_type == "text/html":
                        email_body = BeautifulSoup(part.get_payload(decode=True).decode(), 'html.parser').get_text()
                if email_body:
                    email_body = ' '.join(email_body.split())
                    business_name = extract_business_name(email_body)
                    location = extract_location(email_body).strip()
                    name = re.search(r'Name:(.*?)(?:,|\n)', email_body, re.DOTALL).group(1).strip() if re.search(r'Name:(.*?)(?:,|\n)', email_body, re.DOTALL) else "Name Not Found"
                    phone_number = re.search(r'Phone number:\s*([\d\-\+\(\)\s]+)', email_body).group(1).strip() if re.search(r'Phone number:\s*([\d\-\+\(\)\s]+)', email_body) else "Phone Number Not Found"
                    email_address = re.search(r'Email:\s*([\w\.-]+@[\w\.-]+)', email_body, re.IGNORECASE).group(1).strip() if re.search(r'Email:\s*([\w\.-]+@[\w\.-]+)', email_body, re.IGNORECASE) else "Email Address Not Found"
                    date_sent = datetime.strptime(msg.get("Date"), "%a, %d %b %Y %H:%M:%S %z").strftime("%Y-%m-%d %H:%M:%S %Z") if msg.get("Date") else "Date Not Found"
                    business_link_line = extract_business_link_line(email_body)

                    # Print extracted data
                    print("Business Name:", business_name)
                    print("Location:", location)
                    print("Name:", name)
                    print("Phone Number:", phone_number)
                    print("Email Address:", email_address)
                    print("Date and Time Sent:", date_sent)
                    print("Business Link Line:", business_link_line)

                    # Append extracted data to the main workbook
                    sheet.append([business_name, location, business_link_line, name, phone_number, email_address, date_sent])
        elif "your profile has been suspended" in subject.lower():
            # Process Profile Suspension emails
            content_type = msg.get_content_type()
            if content_type == "multipart/alternative":
                email_body = None
                for part in msg.walk():
                    part_content_type = part.get_content_type()
                    if part_content_type == "text/plain" and email_body is None:
                        email_body = part.get_payload(decode=True).decode()
                    elif part_content_type == "text/html":
                        email_body = BeautifulSoup(part.get_payload(decode=True).decode(), 'html.parser').get_text()
                if email_body:
                    email_body = ' '.join(email_body.split())
                    business_name, business_link_line = extract_suspended_business_info(email_body)

                    # Print extracted data
                    print("Business Name:", business_name)
                    print("Business Link Line:", business_link_line)
                    # Print email content
                    print("Email Content:", email_body)

                    # Add data to the suspended workbook
                    suspended_sheet.append([business_name, business_link_line, email_body])
        elif "Your post has been removed from Google" in subject:
            # Process Removed Posts emails
            content_type = msg.get_content_type()
            if content_type == "multipart/alternative":
                email_body = None
                for part in msg.walk():
                    part_content_type = part.get_content_type()
                    if part_content_type == "text/plain" and email_body is None:
                        email_body = part.get_payload(decode=True).decode()
                    elif part_content_type == "text/html":
                        email_body = BeautifulSoup(part.get_payload(decode=True).decode(), 'html.parser').get_text()
                if email_body:
                    email_body = ' '.join(email_body.split())
                    business_name = extract_removed_post_business_name(email_body)

                    # Print extracted data
                    print("Business Name:", business_name)
                    # Print email content
                    print("Email Content:", email_body)

                    # Add data to the suspended_posts workbook
                    suspended_posts_sheet.append([business_name, email_body])
    except Exception as e:
        print(f"Error processing email ID {email_id}: {str(e)}")

# Save the main workbook
desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
parsed_data_file = os.path.join(desktop_path, "parsed_data.xlsx")
workbook.save(parsed_data_file) 

# Save the suspended workbook as a separate file
suspended_workbook.save(os.path.join(desktop_path, "suspended_gbp_profiles.xlsx"))

# Save the "suspended_posts" workbook as a separate file
suspended_posts_workbook.save(os.path.join(desktop_path, "suspended_posts.xlsx"))

# Logout from the email server
mail.logout()
